"""
Sample college plan — black / tan / white / blue only.
Color on headers only. Alternating tan/white body rows.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# Palette: black, tan, white, blue (one blue for headers only)
BLACK = "111827"
MUTED = "6B7280"
WHITE = "FFFFFF"
TAN = "F5F0E8"
BLUE = "3B82C4"       # header fill
LINE = "E5E0D8"

wb = Workbook()

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
font_sub = Font(name="Calibri", size=10, color=MUTED)
font_sec = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_th = Font(name="Calibri", size=10, bold=True, color=WHITE)
font_body = Font(name="Calibri", size=10, color=BLACK)
font_bold = Font(name="Calibri", size=10, bold=True, color=BLACK)
font_muted = Font(name="Calibri", size=9, italic=True, color=MUTED)

fill_blue = PatternFill("solid", fgColor=BLUE)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_tan = PatternFill("solid", fgColor=TAN)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def widths(ws, lst):
    for i, w in enumerate(lst, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def banner(ws, cols, title, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    c = ws.cell(1, 1, title)
    c.font = font_title
    c.fill = fill_blue
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    for i in range(1, cols + 1):
        ws.cell(1, i).fill = fill_blue

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c = ws.cell(2, 1, subtitle)
    c.font = font_sub
    c.fill = fill_white
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18


def section(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row, 1, text)
    c.font = font_sec
    c.fill = fill_blue
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for i in range(1, cols + 1):
        ws.cell(row, i).fill = fill_blue
    ws.row_dimensions[row].height = 22


def th(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row, i, h)
        cell.font = font_th
        cell.fill = fill_blue
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row].height = 24


def row_data(ws, row, values, done_col=None):
    bg = fill_tan if row % 2 == 0 else fill_white
    for i, v in enumerate(values, 1):
        cell = ws.cell(row, i, v)
        cell.font = font_body
        cell.border = thin
        cell.alignment = left
        cell.fill = bg
    if done_col and done_col <= len(values):
        cell = ws.cell(row, done_col)
        cell.value = "Done"
        cell.font = font_bold
        cell.alignment = center
        # no special color — same row bg
        cell.fill = bg
    if values:
        ws.cell(row, 1).font = font_bold


# ── Sheet 1: Transfer Prep ───────────────────────────────────────────
ws = wb.active
ws.title = "Transfer Prep"
widths(ws, [12, 38, 40, 10, 10, 34])
banner(ws, 6,
       "Transfer prep — IGETC + Economics (Foothill → Berkeley)",
       "AP credit · UC IGETC · ASSIST Econ major prep · Sample: completed path")

ws.merge_cells("A3:F3")
ws["A3"] = (
    "What you get: every pre-transfer box listed next to the course that cleared it. "
    "Built from your transcript, AP list, and ASSIST agreement for your catalog year."
)
ws["A3"].font = font_muted
ws["A3"].alignment = left
ws.row_dimensions[3].height = 28

section(ws, 5, 6, "  AP exams (credit used — scores not listed)")
th(ws, 6, ["Exam", "What it covered", "Where it shows up", "Units*", "Status", "Note"])
for i, r in enumerate([
    ("AP U.S. History", "U.S. history survey", "American History & Institutions", 8, "Done", "Scores omitted"),
    ("AP Physics 1", "Algebra-based physics", "IGETC Area 5A physical science", 8, "Done", "Scores omitted"),
    ("AP Statistics", "Intro statistics", "Quantitative / IGETC Area 2 candidate", 4, "Done", "Not Data C8"),
    ("AP Environmental Science", "Environmental science", "Physical-science AP credit", 4, "Done", "Scores omitted"),
]):
    row_data(ws, 7 + i, list(r), done_col=5)

ws.merge_cells("A11:F11")
ws["A11"] = "* Typical UC-transfer unit values; actual posting follows the UC AP chart."
ws["A11"].font = font_muted

section(ws, 13, 6, "  IGETC — Foothill College (UC pattern)")
th(ws, 14, ["Area", "Requirement", "Completed with", "Units", "Status", "Note"])
igetc = [
    ("1A", "English Composition", "ENGL 1AH", 5, "Done", "Foothill"),
    ("1B", "Critical Thinking / English Comp", "ENGL 1B", 5, "Done", "Foothill"),
    ("2", "Mathematical Concepts", "MATH 1A (also 1B–1D, 2A, 2B)", 5, "Done", "Any Area 2 course clears box"),
    ("3A", "Arts (≥1 of 3 Area-3 courses)", "PHOT 8H", 5, "Done", ""),
    ("3B", "Humanities", "HUMN 1", 4, "Done", ""),
    ("3B", "Humanities (3rd Area-3 course)", "HUMN 7H", 4, "Done", "Completes Area 3 total of 3"),
    ("4", "Social & Behavioral (discipline 1)", "ECON 1A", 5, "Done", "Also Econ prep"),
    ("4", "Social & Behavioral (discipline 1)", "ECON 1B", 5, "Done", ""),
    ("4", "Social & Behavioral (discipline 2)", "PSYC 1", 5, "Done", "Second discipline"),
    ("5A", "Physical Science", "AP Physics 1", 8, "Done", "See AP section"),
    ("5A", "Physical Science (additional)", "AP Environmental Science", 4, "Done", "See AP section"),
    ("5B", "Biological Science", "ANTH 1", 4, "Done", "Foothill"),
    ("6", "Language Other Than English (UC)", "HS LOTE / proficiency", "—", "Done", "Documented"),
    ("—", "IGETC certification", "Full UC IGETC via Foothill", "—", "Done", "Clears L&S breadth + R&C/QR"),
]
for i, r in enumerate(igetc):
    row = 15 + i
    row_data(ws, row, list(r), done_col=5)
    ws.cell(row, 1).alignment = center

ig_end = 14 + len(igetc)
ws.merge_cells(start_row=ig_end + 1, start_column=1, end_row=ig_end + 1, end_column=6)
ws.cell(ig_end + 1, 1, "Full UC IGETC → L&S Seven-Course Breadth + Essential Skills (R&C, QR) via certification.")
ws.cell(ig_end + 1, 1).font = font_bold
ws.cell(ig_end + 1, 1).fill = fill_tan

econ_sec = ig_end + 3
section(ws, econ_sec, 6, "  Economics transfer prep — ASSIST (Foothill → UC Berkeley)")
th(ws, econ_sec + 1, ["Berkeley requirement", "What ASSIST expects", "Completed with", "Units", "Status", "Note"])
for i, r in enumerate([
    ("ECON 1 or 2", "Micro + Macro (usually 2 CCC courses)", "ECON 1A + ECON 1B", "5+5", "Done", "Together = ECON 1"),
    ("MATH 1A or 16A", "First half of year-long calculus", "MATH 1A + MATH 1B", "5+5", "Done", "Quarter artic"),
    ("MATH 1B or 16B", "Second half; same series as A", "MATH 1B + MATH 1C", "5+5", "Done", "Full calc year"),
]):
    row_data(ws, econ_sec + 2 + i, list(r), done_col=5)

ws.freeze_panes = "A7"
ws.sheet_properties.tabColor = BLUE

# ── Sheet 2: Berkeley Graduation ─────────────────────────────────────
ws = wb.create_sheet("Berkeley Graduation")
widths(ws, [28, 36, 36, 10, 10, 32])
banner(ws, 6,
       "Berkeley graduation — Statistics major + L&S rules",
       "CDSS Statistics major · L&S college requirements (first CDSS year) · Graduated sample")

ws.merge_cells("A3:F3")
ws["A3"] = (
    "What you get: lower- and upper-division major boxes plus college/campus rules, "
    "each mapped to the course that satisfied it. This sample used L&S GE rules."
)
ws["A3"].font = font_muted
ws["A3"].alignment = left
ws.row_dimensions[3].height = 28

section(ws, 5, 6, "  Statistics major — lower division")
th(ws, 6, ["Requirement", "Berkeley course", "Completed with", "Units", "Status", "Note"])
for i, r in enumerate([
    ("Calculus I", "Math 1A / 51", "MATH 1A + MATH 1B (Foothill)", "—", "Done", "ASSIST"),
    ("Calculus II", "Math 1B / 52", "MATH 1B + MATH 1C (Foothill)", "—", "Done", "ASSIST"),
    ("Multivariable Calculus", "Math 53", "MATH 1C + MATH 1D (Foothill)", "—", "Done", "ASSIST"),
    ("Linear Algebra", "Math 54 or 56", "MATH 2A + MATH 2B (Foothill)", "—", "Done", "Diff Eq + LA"),
    ("Lower-div Statistics / Data", "Stat 20 or Data C8", "CIS 107 — SJCC (Data 8 curriculum)", "—", "Done", "CIS 107 = Data 8"),
]):
    row_data(ws, 7 + i, list(r), done_col=5)

ws.merge_cells("A13:F13")
ws["A13"] = "Note: CIS 107 (San José City College) follows the UC Berkeley Data 8 curriculum and satisfied Stat 20 / Data C8."
ws["A13"].font = font_muted
ws["A13"].fill = fill_tan

section(ws, 15, 6, "  Statistics major — upper division (9 courses)")
th(ws, 16, ["Category", "Berkeley course", "Completed with", "Units", "Status", "Note"])
for i, r in enumerate([
    ("Core", "STAT 133", "STAT 133", 3, "Done", "Computing with Data"),
    ("Core", "STAT 134", "STAT 134", 4, "Done", "Probability"),
    ("Core", "STAT 135", "STAT 135", 4, "Done", "Concepts of Statistics"),
    ("Stat elective", "Approved UD list", "STAT 154", 4, "Done", "Elective 1"),
    ("Stat elective", "Approved UD list", "STAT 155", 3, "Done", "Elective 2"),
    ("Stat elective", "Approved UD list", "STAT 165", 3, "Done", "Elective 3"),
    ("Cluster", "Applied cluster", "IND ENG 120", 3, "Done", "Cluster 1"),
    ("Cluster", "Applied cluster", "DEMOG C126", 4, "Done", "Cluster 2"),
    ("Cluster", "Applied cluster", "UGBA 102A", 3, "Done", "Cluster 3"),
]):
    row_data(ws, 17 + i, list(r), done_col=5)

section(ws, 28, 6, "  L&S college requirements (used for graduation)")
th(ws, 29, ["Requirement", "How L&S defines it", "Completed with", "Units", "Status", "Note"])
for i, r in enumerate([
    ("Reading & Composition", "2 courses (or IGETC)", "ENGL 1AH + ENGL 1B via IGETC", "5+5", "Done", ""),
    ("Quantitative Reasoning", "1 course (or IGETC)", "MATH series / IGETC Area 2", "—", "Done", ""),
    ("Foreign Language", "L&S language requirement", "HS LOTE / proficiency", "—", "Done", ""),
    ("Seven-Course Breadth", "All 7 areas", "Full UC IGETC certification", "—", "Done", "IGETC = breadth"),
    ("American History & Institutions", "Campus requirement", "AP U.S. History", 8, "Done", "See Transfer Prep"),
    ("American Cultures", "1 approved AC course", "Completed before graduation", "—", "Done", ""),
    ("Unit minimum", "120 semester units", "CC transfer + Berkeley terms", "≥120", "Done", ""),
    ("Upper-division minimum", "36 UD units", "Stats major UD package", "≥36", "Done", ""),
    ("Senior residence", "Final units in residence", "Berkeley enrollment terms", "—", "Done", ""),
    ("GPA floors", "≥2.0 overall & major", "Graduated", "—", "Done", ""),
]):
    row_data(ws, 30 + i, list(r), done_col=5)

ws.freeze_panes = "A7"
ws.sheet_properties.tabColor = BLUE

# ── Sheet 3: Course Plan ─────────────────────────────────────────────
ws = wb.create_sheet("Course Plan")
widths(ws, [10, 14, 34, 8, 16, 50, 10])
banner(ws, 7,
       "Course plan — term by term",
       "Every course in sequence · category · short description · Sample completed path")

ws.merge_cells("A3:G3")
ws["A3"] = (
    "What you get: a living schedule. Category uses a simple dropdown "
    "(Stats Major | L&S / Graduation | Econ Prep | IGETC | AP Credit | Elective | Multiple)."
)
ws["A3"].font = font_muted
ws["A3"].alignment = left
ws.row_dimensions[3].height = 28

th(ws, 5, ["Phase", "Term", "Course", "Units", "Category", "Description", "Status"])
courses = [
    ("Prior", "AP", "AP U.S. History", 8, "AP Credit", "American History & Institutions", "Done"),
    ("Prior", "AP", "AP Physics 1", 8, "AP Credit", "IGETC 5A physical science", "Done"),
    ("Prior", "AP", "AP Statistics", 4, "AP Credit", "Quantitative credit (not Data C8)", "Done"),
    ("Prior", "AP", "AP Environmental Science", 4, "AP Credit", "Physical-science AP credit", "Done"),
    ("CCC", "Spring 2022", "HUMN 1", 4, "IGETC", "IGETC 3B Humanities", "Done"),
    ("CCC", "Spring 2022", "PHOT 8H", 5, "IGETC", "IGETC 3A Arts", "Done"),
    ("CCC", "Summer 2022", "MATH 1A", 5, "Multiple", "IGETC 2 + Stats Calc I + Econ Calc A", "Done"),
    ("CCC", "Summer 2022", "ECON 1A", 5, "Multiple", "IGETC 4 + Econ macro", "Done"),
    ("CCC", "Summer 2022", "ECON 1B", 5, "Multiple", "IGETC 4 + Econ micro", "Done"),
    ("CCC", "Fall 2022", "MATH 1B", 5, "Multiple", "Calc series continuation", "Done"),
    ("CCC", "Fall 2022", "ENGL 1AH", 5, "Multiple", "IGETC 1A + L&S R&C", "Done"),
    ("CCC", "Fall 2022", "C S 1A", 4.5, "Elective", "Intro programming", "Done"),
    ("CCC", "Fall 2022", "HUMN 7H", 4, "IGETC", "IGETC 3B — third Area 3 course", "Done"),
    ("CCC", "Fall 2022", "MUS 12A", 2, "Elective", "Music enrichment", "Done"),
    ("CCC", "Fall 2022", "PSYC 1", 5, "IGETC", "IGETC Area 4 second discipline", "Done"),
    ("CCC", "Winter 2023", "MATH 1C", 5, "Stats Major", "Calc II / multivariable path", "Done"),
    ("CCC", "Winter 2023", "ENGL 1B", 5, "Multiple", "IGETC 1B + L&S R&C", "Done"),
    ("CCC", "Winter 2023", "C S 10", 4.5, "Elective", "CS sequence", "Done"),
    ("CCC", "Winter 2023", "ANTH 1", 4, "IGETC", "IGETC 5B Biological Science", "Done"),
    ("CCC", "Spring 2023", "MATH 1D", 5, "Stats Major", "Math 53 path (with 1C)", "Done"),
    ("CCC", "Spring 2023", "MATH 2B", 5, "Stats Major", "Linear Algebra (Math 54)", "Done"),
    ("CCC", "Spring 2023", "C S 3A", 4.5, "Elective", "CS sequence", "Done"),
    ("CCC", "Fall 2023", "MATH 2A", 5, "Stats Major", "Differential Equations (Math 54)", "Done"),
    ("CCC", "Fall 2023", "PE", 1, "Elective", "Activity units", "Done"),
    ("CCC", "Fall 2023", "CIS 107 (SJCC)", 4.5, "Stats Major", "Data 8 → Stat 20 / Data C8", "Done"),
    ("CCC", "Summer 2024", "ACTG 1A", 5, "Elective", "Accounting elective", "Done"),
    ("CCC", "Summer 2024", "BUS 12", 4, "Elective", "Business elective", "Done"),
    ("UCB", "Spring 2024", "STAT 133", 3, "Stats Major", "UD core — Computing with Data", "Done"),
    ("UCB", "Spring 2024", "STAT 134", 4, "Stats Major", "UD core — Probability", "Done"),
    ("UCB", "Spring 2024", "IND ENG 120", 3, "Stats Major", "UD cluster 1", "Done"),
    ("UCB", "Spring 2024", "PSYC 101", 4, "Elective", "Interest / units", "Done"),
    ("UCB", "Spring 2024", "STAT 198", 1, "Elective", "Special studies", "Done"),
    ("UCB", "Fall 2024", "STAT 135", 4, "Stats Major", "UD core — Concepts of Statistics", "Done"),
    ("UCB", "Fall 2024", "STAT 155", 3, "Stats Major", "UD Stat elective", "Done"),
    ("UCB", "Fall 2024", "DEMOG C126", 4, "Stats Major", "UD cluster 2", "Done"),
    ("UCB", "Spring 2025", "STAT 154", 4, "Stats Major", "UD Stat elective", "Done"),
    ("UCB", "Spring 2025", "STAT 165", 3, "Stats Major", "UD Stat elective", "Done"),
    ("UCB", "Spring 2025", "UGBA 102A", 3, "Stats Major", "UD cluster 3", "Done"),
    ("UCB", "Spring 2025", "LGST 198", 1, "Elective", "Special studies", "Done"),
]
for i, r in enumerate(courses):
    row = 6 + i
    row_data(ws, row, list(r), done_col=7)
    ws.cell(row, 1).alignment = center
    ws.cell(row, 4).alignment = center
    ws.cell(row, 5).alignment = center

last = 5 + len(courses)
dv = DataValidation(
    type="list",
    formula1='"Stats Major,L&S / Graduation,Econ Prep,IGETC,AP Credit,Elective,Multiple"',
    allow_blank=False,
    showDropDown=False,
)
ws.add_data_validation(dv)
dv.add(f"E6:E{last}")
ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A5:G{last}"
ws.sheet_properties.tabColor = BLUE

OUT_DIR = r"C:\Users\amirk\Documents\Grok\blog\assets\sample-packages"
os.makedirs(OUT_DIR, exist_ok=True)
out_xlsx = os.path.join(OUT_DIR, "Amir-College-Plan-Sample.xlsx")
wb.save(out_xlsx)
print("Wrote", out_xlsx)

# ── Screenshots (same palette) ───────────────────────────────────────
from PIL import Image, ImageDraw, ImageFont

SHOT_DIR = os.path.join(OUT_DIR, "screenshots")
os.makedirs(SHOT_DIR, exist_ok=True)

C_BLACK = (17, 24, 39)
C_MUTED = (107, 114, 128)
C_WHITE = (255, 255, 255)
C_TAN = (245, 240, 232)
C_BLUE = (59, 130, 196)
C_LINE = (229, 224, 216)
C_PAGE = (255, 255, 255)


def try_font(size, bold=False):
    paths = (
        [r"C:\Windows\Fonts\segoeuib.ttf", r"C:\Windows\Fonts\calibrib.ttf", r"C:\Windows\Fonts\arialbd.ttf"]
        if bold else
        [r"C:\Windows\Fonts\segoeui.ttf", r"C:\Windows\Fonts\calibri.ttf", r"C:\Windows\Fonts\arial.ttf"]
    )
    for p in paths:
        if os.path.exists(p):
            return ImageFont.truetype(p, size)
    return ImageFont.load_default()


def wrap_text(draw, text, font, max_w):
    text = str(text) if text is not None else ""
    words = text.split() or [""]
    lines, cur = [], words[0]
    for w in words[1:]:
        test = cur + " " + w
        if draw.textlength(test, font=font) <= max_w:
            cur = test
        else:
            lines.append(cur)
            cur = w
    lines.append(cur)
    return lines


def render_table(path, title, subtitle, headers, rows, col_widths, caption=None):
    scale = 2
    pad = 28 * scale
    row_h = 28 * scale
    header_h = 30 * scale
    title_h = 36 * scale
    sub_h = 22 * scale

    font_t = try_font(15 * scale, bold=True)
    font_s = try_font(10 * scale)
    font_h = try_font(10 * scale, bold=True)
    font_c = try_font(10 * scale)
    font_cb = try_font(10 * scale, bold=True)
    font_cap = try_font(9 * scale)

    col_w = [int(w * scale) for w in col_widths]
    table_w = sum(col_w)
    width = table_w + pad * 2

    tmp = Image.new("RGB", (10, 10), C_WHITE)
    dtmp = ImageDraw.Draw(tmp)
    measured = []
    for r in rows:
        cell_lines, max_lines = [], 1
        for i, (cell, cw) in enumerate(zip(r, col_w)):
            lines = wrap_text(dtmp, cell, font_cb if i == 0 else font_c, cw - 16 * scale)
            cell_lines.append(lines)
            max_lines = max(max_lines, len(lines))
        measured.append((cell_lines, max_lines))

    body_h = sum(max(row_h, ml * 16 * scale + 12 * scale) for _, ml in measured)
    cap_h = 36 * scale if caption else 0
    height = pad + title_h + sub_h + 8 * scale + header_h + body_h + cap_h + pad

    img = Image.new("RGB", (width, height), C_PAGE)
    draw = ImageDraw.Draw(img)
    draw.rectangle([pad // 2, pad // 2, width - pad // 2, height - pad // 2], outline=C_LINE, width=scale)

    y = pad
    # Title bar — blue only
    draw.rectangle([pad, y, pad + table_w, y + title_h], fill=C_BLUE)
    draw.text((pad + 12 * scale, y + 8 * scale), title, font=font_t, fill=C_WHITE)
    y += title_h
    draw.text((pad + 12 * scale, y + 2 * scale), subtitle, font=font_s, fill=C_MUTED)
    y += sub_h + 8 * scale

    # Column headers — blue only
    draw.rectangle([pad, y, pad + table_w, y + header_h], fill=C_BLUE)
    x = pad
    for h, cw in zip(headers, col_w):
        draw.text((x + 8 * scale, y + 7 * scale), h, font=font_h, fill=C_WHITE)
        x += cw
    x = pad
    for cw in col_w:
        draw.line([x, y, x, y + header_h], fill=C_LINE, width=scale)
        x += cw
    draw.line([pad + table_w, y, pad + table_w, y + header_h], fill=C_LINE, width=scale)
    draw.line([pad, y + header_h, pad + table_w, y + header_h], fill=C_LINE, width=scale)
    y += header_h

    for idx, (cell_lines, ml) in enumerate(measured):
        rh = max(row_h, ml * 16 * scale + 12 * scale)
        bg = C_TAN if idx % 2 == 1 else C_WHITE
        draw.rectangle([pad, y, pad + table_w, y + rh], fill=bg)
        x = pad
        for i, (lines, cw) in enumerate(zip(cell_lines, col_w)):
            f = font_cb if i == 0 else font_c
            ty = y + 6 * scale
            for line in lines:
                draw.text((x + 8 * scale, ty), line, font=f, fill=C_BLACK)
                ty += 16 * scale
            draw.line([x, y, x, y + rh], fill=C_LINE, width=scale)
            x += cw
        draw.line([pad + table_w, y, pad + table_w, y + rh], fill=C_LINE, width=scale)
        draw.line([pad, y + rh, pad + table_w, y + rh], fill=C_LINE, width=scale)
        y += rh

    if caption:
        draw.text((pad + 4 * scale, y + 10 * scale), caption, font=font_cap, fill=C_MUTED)

    img = img.resize((width // scale, height // scale), Image.Resampling.LANCZOS)
    img.save(path, "PNG", optimize=True)
    print("Shot", path)


render_table(
    os.path.join(SHOT_DIR, "01-transfer-prep-igetc.png"),
    "Transfer Prep — IGETC (sample)",
    "Requirement → course that completed it · Foothill UC pattern",
    ["Area", "Requirement", "Completed with", "Units", "Status"],
    [
        ["1A", "English Composition", "ENGL 1AH", "5", "Done"],
        ["1B", "Critical Thinking / English Comp", "ENGL 1B", "5", "Done"],
        ["2", "Mathematical Concepts", "MATH 1A", "5", "Done"],
        ["3A", "Arts", "PHOT 8H", "5", "Done"],
        ["3B", "Humanities", "HUMN 1", "4", "Done"],
        ["3B", "Humanities (3rd course)", "HUMN 7H", "4", "Done"],
        ["4", "Social & Behavioral", "ECON 1A / 1B + PSYC 1", "—", "Done"],
        ["5A", "Physical Science", "AP Physics 1", "8", "Done"],
        ["5B", "Biological Science", "ANTH 1", "4", "Done"],
        ["6", "LOTE (UC)", "HS LOTE / proficiency", "—", "Done"],
    ],
    [70, 200, 180, 60, 70],
    caption="Excerpt from the Transfer Prep tab. Full sheet also lists AP exams and Econ ASSIST rows.",
)

render_table(
    os.path.join(SHOT_DIR, "02-transfer-prep-econ.png"),
    "Transfer Prep — Economics (ASSIST)",
    "Foothill → UC Berkeley Economics major preparation",
    ["Berkeley requirement", "Completed with", "Units", "Status"],
    [
        ["ECON 1 or 2 (intro)", "ECON 1A + ECON 1B", "5+5", "Done"],
        ["MATH 1A or 16A (Calc A)", "MATH 1A + MATH 1B", "5+5", "Done"],
        ["MATH 1B or 16B (Calc B)", "MATH 1B + MATH 1C", "5+5", "Done"],
    ],
    [220, 200, 80, 70],
    caption="Same tab as IGETC — Econ prep stacked under GE.",
)

render_table(
    os.path.join(SHOT_DIR, "03-berkeley-stats.png"),
    "Berkeley Graduation — Statistics major",
    "Lower division + upper division · sample completed path",
    ["Requirement", "Completed with", "Status"],
    [
        ["Calc I (Math 1A/51)", "MATH 1A + MATH 1B (Foothill)", "Done"],
        ["Calc II (Math 1B/52)", "MATH 1B + MATH 1C (Foothill)", "Done"],
        ["Multivariable (Math 53)", "MATH 1C + MATH 1D (Foothill)", "Done"],
        ["Linear Algebra (Math 54)", "MATH 2A + MATH 2B (Foothill)", "Done"],
        ["Stat 20 / Data C8", "CIS 107 — SJCC (Data 8 curriculum)", "Done"],
        ["UD Core: STAT 133 / 134 / 135", "Taken at Berkeley", "Done"],
        ["UD Electives (3)", "STAT 154, 155, 165", "Done"],
        ["UD Cluster (3)", "IND ENG 120, DEMOG C126, UGBA 102A", "Done"],
    ],
    [240, 280, 70],
    caption="CIS 107 from SJCC maps to Data 8.",
)

render_table(
    os.path.join(SHOT_DIR, "04-berkeley-ls.png"),
    "Berkeley Graduation — L&S college rules",
    "First year of CDSS transition · IGETC does most of the work",
    ["Requirement", "Completed with", "Status"],
    [
        ["Reading & Composition", "ENGL 1AH + ENGL 1B via IGETC", "Done"],
        ["Quantitative Reasoning", "MATH series / IGETC Area 2", "Done"],
        ["Seven-Course Breadth", "Full UC IGETC certification", "Done"],
        ["American History & Institutions", "AP U.S. History", "Done"],
        ["American Cultures", "Completed before graduation", "Done"],
        ["120 units + UD + residence", "CC transfer + Berkeley terms", "Done"],
    ],
    [260, 280, 70],
    caption="L&S essential skills and breadth cleared largely by IGETC certification.",
)

render_table(
    os.path.join(SHOT_DIR, "05-course-plan.png"),
    "Course Plan — term by term (excerpt)",
    "Phase · term · course · category · description",
    ["Term", "Course", "Units", "Category", "Description"],
    [
        ["AP", "AP U.S. History", "8", "AP Credit", "AH&I"],
        ["AP", "AP Physics 1", "8", "AP Credit", "IGETC 5A"],
        ["Summer 2022", "MATH 1A", "5", "Multiple", "IGETC 2 + Stats + Econ calc"],
        ["Summer 2022", "ECON 1A / 1B", "5+5", "Multiple", "IGETC 4 + Econ intro"],
        ["Fall 2022", "ENGL 1AH", "5", "Multiple", "IGETC 1A + R&C"],
        ["Fall 2023", "CIS 107 (SJCC)", "4.5", "Stats Major", "Data 8 → Stat 20 / Data C8"],
        ["Spring 2024", "STAT 133 / 134", "3+4", "Stats Major", "UD core at Berkeley"],
        ["Fall 2024", "STAT 135", "4", "Stats Major", "UD core"],
        ["Spring 2025", "STAT 154 / 165", "4+3", "Stats Major", "UD electives"],
    ],
    [110, 150, 55, 100, 200],
    caption="Full plan lists every dual-enrollment and Berkeley term.",
)

render_table(
    os.path.join(SHOT_DIR, "06-ap-exams.png"),
    "Transfer Prep — AP exams",
    "Scores not listed · credit use only",
    ["Exam", "Where it shows up", "Units", "Status"],
    [
        ["AP U.S. History", "American History & Institutions", "8", "Done"],
        ["AP Physics 1", "IGETC Area 5A physical science", "8", "Done"],
        ["AP Statistics", "Quantitative / IGETC Area 2 candidate", "4", "Done"],
        ["AP Environmental Science", "Physical-science AP credit", "4", "Done"],
    ],
    [200, 260, 70, 70],
    caption="AP list mapped to GE / campus boxes without publishing scores.",
)

print("Done screenshots")
try:
    import shutil
    shutil.copy2(out_xlsx, r"C:\Users\amirk\Downloads\Amir-College-Plan-Sample.xlsx")
except Exception as e:
    print("Downloads copy:", e)
